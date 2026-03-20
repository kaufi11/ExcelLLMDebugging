from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openai import OpenAI
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime
from pydantic import BaseModel

load_dotenv()

client = OpenAI()

class FeedbackEvent(BaseModel):
    faultyCell: str
    repairFormula: str
    errorFound: str
    solutionFound: str
    otherPoints: str

DIR = Path(__file__).resolve().parent
IN_DIR = DIR / "EUSES3/spreadsheets/homework/SEEDED"
IN_PROP_DIR = DIR / "EUSES/configuration_files/homework"
OUT_DIR = DIR / "results_llm_sheet_discription"
OUT_DIR_FOUND = DIR / "results_llm_sheet_discription_found"
DESC_DIR = DIR / "spreadsheet_descriptions"
DESC_DIR.mkdir(exist_ok=True)
OUT_DIR_FOUND.mkdir(exist_ok=True)
OUT_DIR.mkdir(exist_ok=True)

def read_excel_as_text(excel_path: Path) -> str:
    wb = load_workbook(excel_path, data_only=False)
    lines = []

    for sheet in wb.worksheets:
        lines.append(f"Sheet: {sheet.title}")

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        lines.append(f"{cell.coordinate}: FORMULA {cell.value}")
                    else:
                        lines.append(f"{cell.coordinate}: {cell.value}")

    return " | ".join(lines)

def read_txt_as_string(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()
    return content

def describe_spreadsheet_with_llm(excel_text) -> str:
    """
    Uses the LLM to generate a brief description of the spreadsheet's structure and contents.
    """

    response = client.responses.create(
        model="gpt-5-mini",
        temperature=1,
        top_p=1,
        input=[
            {
                "role": "system",
                "content": (
                    "You are an Excel expert. "
                    "Provide a brief description of the spreadsheet structure and contents. "
                    "Summarize sheets, formulas, key values, and overall layout. "
                    "Keep it concise and clear."
                )
            },
            {
                "role": "user",
                "content": excel_text
            }
        ]
    )

    return response.output_text

def analyze_excel(excel_text, discription_text) -> str:

    response = client.responses.create(
        model="gpt-5-mini",
        temperature=1,
        top_p=1,
        input=[
            {
                "role": "system",
                "content": "You are an Excel expert. Analyze the sheet with the help of the discription text which is seperated from the excel_text with a |!| Discription: and find and explain the error/errors and give me "
                "a solution. Also explain the solution. "
                "Give short and easy to understand answers back."
            },
            {
                "role": "user",
                "content": excel_text + " |!| Discription: " + discription_text
            }
        ]
    )

    return response.output_text

def compare_llm_result_with_propertie_file(llm_result, prop_file) -> str:
    propertie_file = read_txt_as_string(prop_file)

    response = client.responses.parse(
        model="gpt-5-mini",
        temperature=1,
        top_p=1,
        input=[
            {
                "role": "system",
                "content": "I asked a llm to find an error. Evaluate with the propertie File if the llm fouond the correct FAULTY_CELL and gave back the right expected value/formula. "
                "Give back only the faulty cell, the repair formula, error found (Yes,No), solution found(Yes,No) and how many other points, which have nothing to do with the FAULTY_CELL as a number otherPoints"
                "The LLM Result and Propertie File are separated by a |"
            },
            {
                "role": "user",
                "content": "LLM Result: " + llm_result + " | Propertie File: " + propertie_file
            }
        ],
        text_format=FeedbackEvent,
    )

    return response.output_parsed


if __name__ == "__main__":

    excel_files = list(IN_DIR.glob("*.xlsx"))

    if not excel_files:
        raise FileNotFoundError(f"No Excel files found in {IN_DIR}")

    for excel_file in excel_files:

        file_name = excel_file.stem
        prop_file = IN_PROP_DIR / f"{file_name}.properties"

        print(f"Analyzing {excel_file.name}...")

        if not prop_file.exists():
            print(f"Property file missing for {excel_file.name}, skipping.")
            continue

        try:
            excel_text = read_excel_as_text(excel_file)
            discription_text = describe_spreadsheet_with_llm(excel_text)
            result_text = analyze_excel(excel_text, discription_text)

            result_found_text = compare_llm_result_with_propertie_file(
                result_text, prop_file
            )

            print(f"Output {result_found_text}")

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            output_file = DESC_DIR / f"{file_name}_analysis_{timestamp}.txt"
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(discription_text)

            output_file = OUT_DIR / f"{file_name}_analysis_{timestamp}.txt"
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(result_text)

            output_found_file = OUT_DIR_FOUND / f"{file_name}_analysis_{timestamp}.txt"
            with open(output_found_file, "w", encoding="utf-8") as f:
                f.write(result_found_text.model_dump_json(indent=2))

            print(f"Result saved to: {output_found_file}")

        except Exception as e:
            print(f"Error processing {excel_file.name}: {e}")
